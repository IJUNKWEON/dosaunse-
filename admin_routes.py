# -*- coding: utf-8 -*-
"""
도사운세 관리자 페이지 라우트
"""

from flask import Blueprint, render_template_string, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_user, logout_user, login_required, current_user
from models import db, Admin, AccessLog, SiteSettings, FortuneCategory, Notice, APIUsage
from datetime import datetime, timedelta
from sqlalchemy import func
import logging
import os
from admin_templates import (
    ADMIN_LOGIN_HTML,
    ADMIN_DASHBOARD_HTML,
    ADMIN_LOGS_HTML,
    ADMIN_SETTINGS_HTML,
    ADMIN_CATEGORIES_HTML,
    ADMIN_ACCOUNT_HTML
)
from admin_templates_ext import (
    # ADMIN_API_USAGE_HTML,  # 숨김 처리
    ADMIN_SYSTEM_HTML,
    ADMIN_NOTICES_HTML
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

logger = logging.getLogger(__name__)

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

# 접속 로그 기록 함수
def log_access(fortune_type, user_data=None):
    """사용자 접속 기록"""
    try:
        log = AccessLog(
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent', '')[:255],
            fortune_type=fortune_type,
            birth_date=user_data.get('birthDate') if user_data else None,
            birth_time=user_data.get('birthTime') if user_data else None,
            gender=user_data.get('gender') if user_data else None,
            calendar_type=user_data.get('calendarType') if user_data else None
        )
        db.session.add(log)
        db.session.commit()
    except Exception as e:
        logger.error(f"접속 로그 기록 실패: {e}")
        db.session.rollback()


@admin_bp.route('/login', methods=['GET', 'POST'])
def admin_login():
    """관리자 로그인"""
    if current_user.is_authenticated:
        return redirect(url_for('admin.dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        admin = Admin.query.filter_by(username=username).first()
        
        if admin and admin.check_password(password):
            login_user(admin)
            admin.last_login = datetime.utcnow()
            db.session.commit()
            logger.info(f"관리자 로그인: {username}")
            return redirect(url_for('admin.dashboard'))
        else:
            flash('아이디 또는 비밀번호가 잘못되었습니다.', 'error')
    
    # 로그인 페이지 HTML
    return render_template_string(ADMIN_LOGIN_HTML)


@admin_bp.route('/logout')
@login_required
def admin_logout():
    """관리자 로그아웃"""
    logger.info(f"관리자 로그아웃: {current_user.username}")
    logout_user()
    return redirect(url_for('admin.admin_login'))


@admin_bp.route('/')
@admin_bp.route('/dashboard')
@login_required
def dashboard():
    """관리자 대시보드"""
    # 통계 데이터 계산
    today = datetime.now().date()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    
    # 전체 접속 수
    total_access = AccessLog.query.count()
    today_access = AccessLog.query.filter(
        func.date(AccessLog.timestamp) == today
    ).count()
    week_access = AccessLog.query.filter(
        AccessLog.timestamp >= datetime.combine(week_ago, datetime.min.time())
    ).count()
    
    # 운세별 통계
    fortune_stats = db.session.query(
        AccessLog.fortune_type,
        func.count(AccessLog.id).label('count')
    ).group_by(AccessLog.fortune_type).order_by(func.count(AccessLog.id).desc()).limit(10).all()
    
    # 성별 통계
    gender_stats = db.session.query(
        AccessLog.gender,
        func.count(AccessLog.id).label('count')
    ).filter(AccessLog.gender.isnot(None)).group_by(AccessLog.gender).all()
    
    # 최근 접속 로그
    recent_logs = AccessLog.query.order_by(AccessLog.timestamp.desc()).limit(20).all()
    
    # 시간대별 접속 통계 (오늘)
    hourly_stats = db.session.query(
        func.strftime('%H', AccessLog.timestamp).label('hour'),
        func.count(AccessLog.id).label('count')
    ).filter(
        func.date(AccessLog.timestamp) == today
    ).group_by('hour').all()
    
    stats = {
        'total_access': total_access,
        'today_access': today_access,
        'week_access': week_access,
        'fortune_stats': fortune_stats,
        'gender_stats': gender_stats,
        'recent_logs': recent_logs,
        'hourly_stats': hourly_stats
    }
    
    return render_template_string(ADMIN_DASHBOARD_HTML, stats=stats, admin=current_user)


@admin_bp.route('/logs')
@login_required
def logs():
    """접속 로그 조회"""
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    # 필터링
    fortune_type = request.args.get('fortune_type')
    gender = request.args.get('gender')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    query = AccessLog.query
    
    if fortune_type:
        query = query.filter(AccessLog.fortune_type == fortune_type)
    if gender:
        query = query.filter(AccessLog.gender == gender)
    if date_from:
        query = query.filter(AccessLog.timestamp >= datetime.strptime(date_from, '%Y-%m-%d'))
    if date_to:
        query = query.filter(AccessLog.timestamp <= datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
    
    pagination = query.order_by(AccessLog.timestamp.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return render_template_string(ADMIN_LOGS_HTML, 
                                 pagination=pagination,
                                 fortune_type=fortune_type,
                                 gender=gender,
                                 date_from=date_from,
                                 date_to=date_to,
                                 admin=current_user)


@admin_bp.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    """사이트 설정 관리"""
    if request.method == 'POST':
        # 설정 업데이트
        for key, value in request.form.items():
            if key.startswith('setting_'):
                setting_key = key.replace('setting_', '')
                setting = SiteSettings.query.filter_by(key=setting_key).first()
                if setting:
                    setting.value = value
                    setting.updated_at = datetime.utcnow()
                else:
                    setting = SiteSettings(key=setting_key, value=value)
                    db.session.add(setting)
        
        db.session.commit()
        flash('설정이 저장되었습니다.', 'success')
        return redirect(url_for('admin.settings'))
    
    # 모든 설정 조회
    all_settings = SiteSettings.query.all()
    
    return render_template_string(ADMIN_SETTINGS_HTML, 
                                 settings=all_settings,
                                 admin=current_user)


@admin_bp.route('/categories', methods=['GET', 'POST'])
@login_required
def categories():
    """운세 카테고리 관리"""
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'add':
            category = FortuneCategory(
                name=request.form.get('name'),
                key=request.form.get('key'),
                icon=request.form.get('icon'),
                color=request.form.get('color'),
                description=request.form.get('description'),
                is_active=request.form.get('is_active') == 'on',
                sort_order=int(request.form.get('sort_order', 0))
            )
            db.session.add(category)
            db.session.commit()
            flash('카테고리가 추가되었습니다.', 'success')
        
        elif action == 'update':
            category_id = request.form.get('category_id')
            category = FortuneCategory.query.get(category_id)
            if category:
                category.name = request.form.get('name')
                category.icon = request.form.get('icon')
                category.color = request.form.get('color')
                category.description = request.form.get('description')
                category.is_active = request.form.get('is_active') == 'on'
                category.sort_order = int(request.form.get('sort_order', 0))
                db.session.commit()
                flash('카테고리가 수정되었습니다.', 'success')
        
        elif action == 'delete':
            category_id = request.form.get('category_id')
            category = FortuneCategory.query.get(category_id)
            if category:
                db.session.delete(category)
                db.session.commit()
                flash('카테고리가 삭제되었습니다.', 'success')
        
        return redirect(url_for('admin.categories'))
    
    # 모든 카테고리 조회
    all_categories = FortuneCategory.query.order_by(FortuneCategory.sort_order).all()
    
    return render_template_string(ADMIN_CATEGORIES_HTML, 
                                 categories=all_categories,
                                 admin=current_user)


@admin_bp.route('/logs/export', methods=['GET'])
@login_required
def export_logs():
    """접속 로그 엑셀 다운로드"""
    try:
        # 필터링 파라미터
        fortune_type = request.args.get('fortune_type')
        gender = request.args.get('gender')
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        
        # 쿼리 생성
        query = AccessLog.query
        
        if fortune_type:
            query = query.filter(AccessLog.fortune_type == fortune_type)
        if gender:
            query = query.filter(AccessLog.gender == gender)
        if date_from:
            query = query.filter(AccessLog.timestamp >= datetime.strptime(date_from, '%Y-%m-%d'))
        if date_to:
            query = query.filter(AccessLog.timestamp <= datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
        
        logs = query.order_by(AccessLog.timestamp.desc()).all()
        
        # 엑셀 파일 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "접속 로그"
        
        # 헤더 스타일 설정
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 헤더 작성
        headers = ["번호", "접속시간", "IP주소", "운세종류", "생년월일", "출생시간", "성별", "양/음력"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 데이터 작성
        for idx, log in enumerate(logs, 2):
            ws.cell(row=idx, column=1, value=idx-1)
            ws.cell(row=idx, column=2, value=log.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=idx, column=3, value=log.ip_address or '-')
            ws.cell(row=idx, column=4, value=log.fortune_type or '-')
            ws.cell(row=idx, column=5, value=log.birth_date or '-')
            ws.cell(row=idx, column=6, value=log.birth_time or '-')
            ws.cell(row=idx, column=7, value=log.gender or '-')
            ws.cell(row=idx, column=8, value=log.calendar_type or '-')
        
        # 열 너비 자동 조정
        column_widths = [8, 20, 15, 15, 15, 15, 10, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # 메모리에 저장
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # 파일명 생성
        filename = f"도사운세_접속로그_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        logger.info(f"엑셀 다운로드: {current_user.username}, {len(logs)}건")
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        logger.error(f"엑셀 다운로드 실패: {e}")
        flash('엑셀 다운로드 중 오류가 발생했습니다.', 'error')
        return redirect(url_for('admin.logs'))


@admin_bp.route('/account', methods=['GET', 'POST'])
@login_required
def account():
    """관리자 계정 관리"""
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'change_password':
            current_password = request.form.get('current_password')
            new_password = request.form.get('new_password')
            confirm_password = request.form.get('confirm_password')
            
            # 현재 비밀번호 확인
            if not current_user.check_password(current_password):
                flash('현재 비밀번호가 올바르지 않습니다.', 'error')
                return redirect(url_for('admin.account'))
            
            # 새 비밀번호 확인
            if new_password != confirm_password:
                flash('새 비밀번호가 일치하지 않습니다.', 'error')
                return redirect(url_for('admin.account'))
            
            if len(new_password) < 6:
                flash('비밀번호는 최소 6자 이상이어야 합니다.', 'error')
                return redirect(url_for('admin.account'))
            
            # 비밀번호 변경
            current_user.set_password(new_password)
            db.session.commit()
            flash('비밀번호가 변경되었습니다.', 'success')
            logger.info(f"비밀번호 변경: {current_user.username}")
            return redirect(url_for('admin.account'))
        
        elif action == 'add_admin':
            username = request.form.get('username')
            email = request.form.get('email')
            password = request.form.get('password')
            
            # 중복 확인
            if Admin.query.filter_by(username=username).first():
                flash('이미 존재하는 아이디입니다.', 'error')
                return redirect(url_for('admin.account'))
            
            if Admin.query.filter_by(email=email).first():
                flash('이미 존재하는 이메일입니다.', 'error')
                return redirect(url_for('admin.account'))
            
            # 새 관리자 추가
            new_admin = Admin(username=username, email=email)
            new_admin.set_password(password)
            db.session.add(new_admin)
            db.session.commit()
            flash('새 관리자 계정이 추가되었습니다.', 'success')
            logger.info(f"새 관리자 추가: {username} by {current_user.username}")
            return redirect(url_for('admin.account'))
        
        elif action == 'delete_admin':
            admin_id = request.form.get('admin_id')
            admin_to_delete = Admin.query.get(admin_id)
            
            if admin_to_delete:
                if admin_to_delete.id == current_user.id:
                    flash('자기 자신을 삭제할 수 없습니다.', 'error')
                else:
                    username = admin_to_delete.username
                    db.session.delete(admin_to_delete)
                    db.session.commit()
                    flash(f'{username} 계정이 삭제되었습니다.', 'success')
                    logger.info(f"관리자 삭제: {username} by {current_user.username}")
            
            return redirect(url_for('admin.account'))
    
    # 모든 관리자 조회
    all_admins = Admin.query.order_by(Admin.created_at).all()
    
    return render_template_string(ADMIN_ACCOUNT_HTML, 
                                 admins=all_admins,
                                 admin=current_user)


# API 사용량 페이지는 숨김 처리 (내부적으로는 계속 추적)
# @admin_bp.route('/api-usage')
# @login_required
# def api_usage():
#     """API 사용량 모니터링"""
#     # 기간별 통계
#     today = datetime.now().date()
#     week_ago = today - timedelta(days=7)
#     month_ago = today - timedelta(days=30)
#     
#     # 오늘 사용량
#     today_usage = db.session.query(
#         func.sum(APIUsage.tokens_used).label('tokens'),
#         func.sum(APIUsage.estimated_cost).label('cost')
#     ).filter(func.date(APIUsage.timestamp) == today).first()
#     
#     # 이번 주 사용량
#     week_usage = db.session.query(
#         func.sum(APIUsage.tokens_used).label('tokens'),
#         func.sum(APIUsage.estimated_cost).label('cost')
#     ).filter(APIUsage.timestamp >= datetime.combine(week_ago, datetime.min.time())).first()
#     
#     # 이번 달 사용량
#     month_usage = db.session.query(
#         func.sum(APIUsage.tokens_used).label('tokens'),
#         func.sum(APIUsage.estimated_cost).label('cost')
#     ).filter(APIUsage.timestamp >= datetime.combine(month_ago, datetime.min.time())).first()
#     
#     # 카테고리별 사용량
#     category_usage = db.session.query(
#         APIUsage.category,
#         func.count(APIUsage.id).label('count'),
#         func.sum(APIUsage.tokens_used).label('tokens'),
#         func.sum(APIUsage.estimated_cost).label('cost')
#     ).group_by(APIUsage.category).order_by(func.sum(APIUsage.estimated_cost).desc()).all()
#     
#     # 최근 사용 기록
#     recent_usage = APIUsage.query.order_by(APIUsage.timestamp.desc()).limit(20).all()
#     
#     stats = {
#         'today': today_usage,
#         'week': week_usage,
#         'month': month_usage,
#         'category_usage': category_usage,
#         'recent_usage': recent_usage
#     }
#     
#     return render_template_string(ADMIN_API_USAGE_HTML, stats=stats, admin=current_user)


@admin_bp.route('/system')
@login_required
def system_status():
    """도사운세 데이터베이스 통계"""
    # 데이터베이스 크기
    db_path = 'dosa_admin.db'
    db_size = os.path.getsize(db_path) / (1024 * 1024) if os.path.exists(db_path) else 0  # MB
    
    # 데이터베이스 통계
    total_logs = AccessLog.query.count()
    total_api_calls = APIUsage.query.count()
    total_admins = Admin.query.count()
    total_notices = Notice.query.count()
    
    # 최근 7일간 일별 통계
    today = datetime.now().date()
    week_ago = today - timedelta(days=7)
    
    daily_stats = db.session.query(
        func.date(AccessLog.timestamp).label('date'),
        func.count(AccessLog.id).label('count')
    ).filter(
        AccessLog.timestamp >= datetime.combine(week_ago, datetime.min.time())
    ).group_by(func.date(AccessLog.timestamp)).all()
    
    # 성별 통계
    gender_stats = db.session.query(
        AccessLog.gender,
        func.count(AccessLog.id).label('count')
    ).group_by(AccessLog.gender).all()
    
    # 운세 카테고리별 통계
    category_stats = db.session.query(
        AccessLog.fortune_type,
        func.count(AccessLog.id).label('count')
    ).group_by(AccessLog.fortune_type).order_by(func.count(AccessLog.id).desc()).limit(10).all()
    
    system_info = {
        'db_size': round(db_size, 2),
        'total_logs': total_logs,
        'total_api_calls': total_api_calls,
        'total_admins': total_admins,
        'total_notices': total_notices,
        'daily_stats': daily_stats,
        'gender_stats': gender_stats,
        'category_stats': category_stats
    }
    
    return render_template_string(ADMIN_SYSTEM_HTML, system_info=system_info, admin=current_user)


@admin_bp.route('/notices', methods=['GET', 'POST'])
@login_required
def notices():
    """공지사항 관리"""
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'add':
            notice = Notice(
                title=request.form.get('title'),
                content=request.form.get('content'),
                is_active=request.form.get('is_active') == 'on',
                priority=int(request.form.get('priority', 0))
            )
            db.session.add(notice)
            db.session.commit()
            flash('공지사항이 추가되었습니다.', 'success')
            logger.info(f"공지사항 추가: {notice.title} by {current_user.username}")
        
        elif action == 'update':
            notice_id = request.form.get('notice_id')
            notice = Notice.query.get(notice_id)
            if notice:
                notice.title = request.form.get('title')
                notice.content = request.form.get('content')
                notice.is_active = request.form.get('is_active') == 'on'
                notice.priority = int(request.form.get('priority', 0))
                notice.updated_at = datetime.utcnow()
                db.session.commit()
                flash('공지사항이 수정되었습니다.', 'success')
        
        elif action == 'delete':
            notice_id = request.form.get('notice_id')
            notice = Notice.query.get(notice_id)
            if notice:
                db.session.delete(notice)
                db.session.commit()
                flash('공지사항이 삭제되었습니다.', 'success')
        
        return redirect(url_for('admin.notices'))
    
    # 모든 공지사항 조회
    all_notices = Notice.query.order_by(Notice.priority.desc(), Notice.created_at.desc()).all()
    
    return render_template_string(ADMIN_NOTICES_HTML, notices=all_notices, admin=current_user)

